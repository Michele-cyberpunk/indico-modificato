# This file is part of the ECM extension for Indico.
# Copyright (C) 2026 - present
#
# Licensed under the MIT License; see the LICENSE file for details.

"""The guest list as the spreadsheet the office already passes around.

Ported from `src/lib/transfer/transferExcelExport.ts` of the Cyberbrain event
manager: two sheets, `Andata` and `Ritorno`, the same columns in the same
order, and the four counters at the bottom. The office forwards this file to
the hotel and to the shuttle company, so the shape is not ours to improve.

The itinerary — who is picked up with whom, in which vehicle, at which time —
is the grouping `guests.group_transfers` already computes; here it is written
out per run so a driver reads one page per trip.

Pure functions, no Indico imports.
"""

import io


#: Columns of both sheets, in the order the office reads them
COLUMNS = ('Cognome', 'Nome', 'Città', 'Mezzo', 'Orario Arrivo', 'Orario Partenza',
           'Cena', 'Note', 'Data andata', 'Data ritorno')

SHEET_OUTBOUND = 'Andata'
SHEET_RETURN = 'Ritorno'
SHEET_ITINERARY = 'Itinerari'

#: The counters written under the list
SUMMARY_LABELS = ('OSPITI TOTALI', 'TRASPORTO PROPRIO', 'CENE RICHIESTE', 'CON PERNOTTAMENTO')


def _time(value):
    return value.strftime('%H:%M') if value else ''


def _yes_no(value):
    return 'Sì' if value else ''


def guest_row(guest, *, outbound_date='', return_date=''):
    """One guest, as the row the office expects."""
    return (
        guest.last_name,
        guest.first_name,
        guest.transfer_place,
        'Proprio' if guest.own_transport else 'Navetta',
        _time(guest.arrival),
        _time(guest.departure),
        _yes_no(guest.dinner),
        guest.diet_notes or guest.notes or '',
        outbound_date,
        return_date,
    )


def summary(guests):
    """The four counters the original writes under the list.

    `CON PERNOTTAMENTO` counts whoever has both an arrival and a departure: they
    are staying the night between the two, which is what the hotel is booked for.
    """
    guests = list(guests)
    return {
        'OSPITI TOTALI': sum(guest.pax for guest in guests),
        'TRASPORTO PROPRIO': sum(1 for guest in guests if guest.own_transport),
        'CENE RICHIESTE': sum(guest.pax for guest in guests if guest.dinner),
        'CON PERNOTTAMENTO': sum(1 for guest in guests if guest.arrival and guest.departure),
    }


def itinerary_rows(groups, *, arrival=True):
    """One block per run: the window, the vehicle, then who is on board."""
    rows = []
    heading = 'Arrivo' if arrival else 'Partenza'
    for group in groups:
        vehicle = f' — veicolo {group.vehicle_number}' if getattr(group, 'vehicle_number', None) else ''
        rows.append((f'{heading} {group.window}{vehicle}', '', '', ''))
        rows.append(('Nome', 'Telefono', 'Pax', 'Orario'))
        for guest in group.guests:
            moment = guest.arrival if arrival else guest.departure
            rows.append((guest.full_name, guest.phone or '', guest.pax, _time(moment)))
        rows.append(('', '', '', ''))
    return rows


def _write_sheet(sheet, guests, *, outbound_date, return_date):
    sheet.append(list(COLUMNS))
    for guest in guests:
        sheet.append(list(guest_row(guest, outbound_date=outbound_date, return_date=return_date)))
    sheet.append([])
    for label, value in summary(guests).items():
        sheet.append([label, value])


def build_workbook(guests, *, arrivals=(), departures=(), event_date=None, return_date=None):
    """The whole file: outbound, return, and the itineraries.

    `arrivals` and `departures` are what `guests.group_transfers` returned; pass
    them and the third sheet carries one block per run.
    """
    from openpyxl import Workbook

    guests = list(guests)
    outbound = event_date.strftime('%d/%m/%Y') if event_date else ''
    back = return_date.strftime('%d/%m/%Y') if return_date else outbound

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_OUTBOUND
    _write_sheet(sheet, [guest for guest in guests if not guest.own_transport],
                 outbound_date=outbound, return_date=back)

    returning = workbook.create_sheet(SHEET_RETURN)
    _write_sheet(returning, [guest for guest in guests if guest.departure],
                 outbound_date=outbound, return_date=back)

    if arrivals or departures:
        itinerary = workbook.create_sheet(SHEET_ITINERARY)
        for row in itinerary_rows(arrivals, arrival=True):
            itinerary.append(list(row))
        for row in itinerary_rows(departures, arrival=False):
            itinerary.append(list(row))

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def filename_for(event_id):
    return f'transfer-{event_id}.xlsx'
