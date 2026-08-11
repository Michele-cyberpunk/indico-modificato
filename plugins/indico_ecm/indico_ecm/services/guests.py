# This file is part of the ECM extension for Indico.
# Copyright (C) 2026 - present
#
# Licensed under the MIT License; see the LICENSE file for details.

"""The guest list: from a messy sponsor list to transfers and covers.

This is the second half of Cyberbrain's document intake, and the half that never
needed a model at all. The legacy app tried Gemini first, cascading across three
models, and fell back to `extractWithRegex` when the key was missing or every
quota was exhausted. Here the rules come first and are the only path that
exists: a sponsor's participant list is short, formulaic text, and a regular
expression that has been tested is more predictable — and auditable — than a
model that answers differently on Tuesday.

What it does, in order:

1. `extract_guest` reads one line into a `Guest`.
2. `import_guest_list` turns a whole file into guests and *rejected rows with a
   reason*, because a row that is silently dropped is a person who arrives at an
   airport with nobody waiting.
3. `categorize` splits them into arrivals, departures, own transport, no
   transfer, lunch and dinner.
4. `group_transfers` packs them into time windows and vehicles.

Pure, no Indico imports.
"""

import re
import unicodedata
from dataclasses import dataclass, field, replace
from datetime import time


# --- reading one row ----------------------------------------------------------------

EMAIL_RE = re.compile(r'[\w.+-]+@[\w-]+(?:\.[\w-]+)+')

#: An Italian mobile or landline, written the way people write it. Anchored on a
#: boundary that is not a digit-adjacent separator so it cannot start halfway
#: through a longer number, and required to hold 8-13 digits in total.
PHONE_RE = re.compile(r'(?<![\d/-])(?:\+\d{1,3}[\s.-]?)?(?:\(?\d{2,4}\)?[\s.-]?){2,4}\d{2,4}(?![\d/-])')

#: `h` is a real Italian abbreviation for a time, but unanchored it also matches
#: the last letter of a surname: `Bosch: 10:30` used to be read as an arrival.
_ARRIVAL_WORDS = r'arrivo|arr\.|\barr\b|\bh\b|atterraggio|check-?in'
_DEPARTURE_WORDS = r'partenza|part\.|\bpart\b|\bdep\b|ripartenza|check-?out|rientro'
_TIME = r'(\d{1,2})[:.](\d{2})'
ARRIVAL_RE = re.compile(rf'(?:{_ARRIVAL_WORDS})\s*:?\s*{_TIME}', re.IGNORECASE)
DEPARTURE_RE = re.compile(rf'(?:{_DEPARTURE_WORDS})\s*:?\s*{_TIME}', re.IGNORECASE)

#: "+ 1 accompagnatore", "2 pax", "x2". The legacy code hardcoded pax to 1, so a
#: doctor travelling with a companion silently lost a seat on the shuttle.
PAX_RE = re.compile(r'(?:\+\s*(\d)\s*(?:accompagnator[ei]|persona|persone|pax)?'
                    r'|(\d{1,2})\s*pax\b'
                    r'|\bpax\s*:?\s*(\d{1,2})\b'
                    r'|\bx\s*(\d)\b)', re.IGNORECASE)

OWN_TRANSPORT_RE = re.compile(r'\b(?:auto\s*propria|mezzi\s*propri|mezzo\s*proprio|trasporto\s*proprio'
                              r'|autonomo|in\s*auto|own\s*car|self\s*drive|noleggio|rental)\b', re.IGNORECASE)

#: A meal is requested unless it is denied, and the denial can come before the
#: word as easily as after it: "no pranzo" used to count as a lunch.
_MEAL_NO = r'(?:no|nessun[oa]?|non|senza|x|-)'


def _meal_requested(text, *words):
    """Whether a meal is asked for, reading the denial on either side."""
    joined = '|'.join(words)
    if re.search(rf'\b{_MEAL_NO}\s+(?:{joined})\b', text, re.IGNORECASE):
        return False
    if re.search(rf'\b(?:{joined})\s*:?\s*{_MEAL_NO}\b', text, re.IGNORECASE):
        return False
    return bool(re.search(rf'\b(?:{joined})\b', text, re.IGNORECASE))


#: Diets that are the value themselves: "vegetariano" says everything, and the
#: legacy pattern threw it away because it only kept what came *after* the word.
STANDALONE_DIETS = ('vegetariano', 'vegetariana', 'vegano', 'vegana', 'celiaco', 'celiaca',
                    'senza glutine', 'senza lattosio', 'gluten free', 'lactose free')
DIET_RE = re.compile(r'(?:allergi[ae]|intolleranz[ae]|dieta|allergico\s+a|intollerante\s+a)'
                     r'\s*:?\s*([^,;\n]+)', re.IGNORECASE)

_TITLE_RE = re.compile(r'\b(?:Dott\.ssa|Dott\.|Dot\.|Prof\.ssa|Prof\.|Dr\.ssa|Dr\.|Sig\.ra|Sig\.|Ing\.)\s*',
                       re.IGNORECASE)

#: Surname particles, lowercase in Italian and kept as part of the name
_PARTICLE = r"(?:de(?:gli|lla|llo|lle|l|i|n|r)?|di|da|dal|del|della|van|von|ten|ter|la|lo|d')"
_WORD = r"[A-ZÀ-Ú][\w'\u2019À-ú]*|[A-ZÀ-Ú]{2,}"
NAME_TOKEN_RE = re.compile(rf'(?:{_WORD}|{_PARTICLE})(?![\w@])')

#: Words that appear in a participant list but never in a person's name. A row
#: of logistics ("Hotel Excelsior Milano") used to become a guest called Hotel.
NOT_A_NAME = frozenset({
    'hotel', 'albergo', 'sala', 'aula', 'via', 'viale', 'piazza', 'corso', 'centro',
    'ospedale', 'clinica', 'azienda', 'asl', 'aou', 'irccs', 'policlinico', 'istituto',
    'congressi', 'congresso', 'evento', 'transfer', 'navetta', 'volo', 'treno', 'stazione',
    'aeroporto', 'arrivo', 'arrivi', 'partenza', 'partenze', 'pranzo', 'cena', 'note',
    'nome', 'cognome', 'email', 'mail', 'telefono', 'cell', 'cellulare', 'pax', 'totale',
    'reparto', 'ruolo', 'relatore', 'moderatore', 'presidente', 'accompagnatore',
})


@dataclass
class Guest:
    """One person on the list, and what the office has to arrange for them."""

    first_name: str = ''
    last_name: str = ''
    email: str = ''
    phone: str = ''
    pax: int = 1
    arrival: time | None = None
    departure: time | None = None
    transfer_place: str = ''
    own_transport: bool = False
    lunch: bool = False
    dinner: bool = False
    diet_notes: str = ''
    notes: str = ''
    row_number: int = 0
    #: False when nothing in the row said which token was the surname
    name_order_certain: bool = True
    #: field -> how it was read, so a wrong value can be traced to its rule
    evidence: dict = field(default_factory=dict)

    @property
    def full_name(self):
        return ' '.join(part for part in (self.first_name, self.last_name) if part)

    def swapped(self):
        """The same guest with given name and surname exchanged."""
        return replace(self, first_name=self.last_name, last_name=self.first_name,
                       name_order_certain=True)

    @property
    def has_contact(self):
        return bool(self.email or self.phone)

    @property
    def needs_transfer(self):
        return not self.own_transport and (self.arrival is not None or self.departure is not None)


@dataclass(frozen=True)
class RejectedRow:
    """A row that produced no usable guest, and why.

    Kept and shown rather than dropped: a silently discarded row is a person who
    lands at the airport with nobody waiting for them.
    """

    row_number: int
    content: str
    reason: str


def _strip_accents(text):
    return ''.join(char for char in unicodedata.normalize('NFD', text)
                   if unicodedata.category(char) != 'Mn')


def _parse_time(hours, minutes):
    hours, minutes = int(hours), int(minutes)
    if 0 <= hours <= 23 and 0 <= minutes <= 59:
        return time(hours, minutes)
    return None


def _search_time(pattern, text):
    match = pattern.search(text)
    if not match:
        return None, ''
    return _parse_time(match.group(1), match.group(2)), match.group(0)


def find_pax(text):
    """How many people travel under this name."""
    match = PAX_RE.search(text)
    if not match:
        return 1, ''
    value = next((group for group in match.groups() if group), None)
    if value is None:
        return 1, ''
    count = int(value)
    # "+ 1 accompagnatore" is two people; "2 pax" is already the total
    if match.group(1):
        count += 1
    return (count if 1 <= count <= 20 else 1), match.group(0)


#: Given names common enough in Italian participant lists to tell `Rossi Mario`
#: from `Mario Rossi`. Deliberately a short list of frequent names: it decides
#: the common case and stays silent — rather than wrong — on the rest.
GIVEN_NAMES = frozenset((
    'alberto', 'aldo', 'alessandra', 'alessandro', 'alessia', 'alessio', 'andrea', 'angelo', 'anna',
    'annamaria', 'antonella', 'antonio', 'arianna', 'armando', 'barbara', 'beatrice', 'benedetta',
    'bianca', 'bruno', 'camilla', 'carla', 'carlo', 'carmela', 'caterina', 'cecilia', 'chiara',
    'christian', 'claudia', 'claudio', 'cristina', 'daniela', 'daniele', 'dario', 'davide', 'diego',
    'domenico', 'elena', 'eleonora', 'elisa', 'elisabetta', 'emanuela', 'emanuele', 'enrico',
    'eugenio', 'fabio', 'fabrizio', 'federica', 'federico', 'filippo', 'flavia', 'francesca',
    'francesco', 'franco', 'gabriele', 'gaetano', 'gaia', 'giacomo', 'gianluca', 'gianni', 'gino',
    'giorgia', 'giorgio', 'giovanna', 'giovanni', 'giulia', 'giulio', 'giuseppe', 'grazia', 'guido',
    'ignazio', 'ilaria', 'irene', 'isabella', 'ivan', 'jacopo', 'jessica', 'laura', 'leonardo',
    'lorenzo', 'luca', 'lucia', 'luciana', 'luciano', 'ludovica', 'luigi', 'luisa', 'manuela',
    'marcello', 'marco', 'maria', 'mariagrazia', 'marina', 'mario', 'marta', 'martina', 'massimo',
    'matteo', 'mattia', 'maurizio', 'michela', 'michele', 'milena', 'mirko', 'monica', 'nadia',
    'niccolo', 'nicola', 'nicoletta', 'nicolo', 'ornella', 'paola', 'paolo', 'patrizia',
    'pierluigi', 'pietro', 'raffaele', 'raffaella', 'renato', 'riccardo', 'rita', 'roberta',
    'roberto', 'rosa', 'rossella', 'sabrina', 'salvatore', 'samuele', 'sandra', 'sara', 'serena',
    'sergio', 'silvana', 'silvia', 'simona', 'simone', 'sofia', 'sonia', 'stefania', 'stefano',
    'teresa', 'tiziana', 'tiziano', 'tommaso', 'umberto', 'valentina', 'valeria', 'valerio',
    'vanessa', 'vincenzo', 'vittoria', 'walter'
))


@dataclass(frozen=True)
class Name:
    """A name, and how sure the rule is that it read the order right."""

    first: str = ''
    last: str = ''
    evidence: str = ''
    #: False when nothing in the row said which token is the surname
    certain: bool = True


def _pretty(token):
    """A SHOUTED token written as a name; anything else is left alone."""
    return token.title() if token.isupper() and len(token) > 1 else token


def _split_name(tokens, *, surname_first, certain=True):
    evidence = ' '.join(tokens)
    shown = [_pretty(token) for token in tokens]
    if surname_first:
        return Name(shown[-1], ' '.join(shown[:-1]), evidence, certain)
    return Name(shown[0], ' '.join(shown[1:]), evidence, certain)


def find_name(text):
    """The person's name, without the parts that are not a name.

    `Mario Rossi` and `Rossi Mario` are the same characters in the same order:
    no rule can separate them, so instead of guessing this looks for something
    in the row that actually says which is which — a comma (`Rossi, Mario`), a
    surname written in capitals (`ROSSI Mario`), or a recognisable given name —
    and when none of them is there it says so with `certain=False` rather than
    pretending. The page lets a person swap the two in one click.
    """
    without_contacts = EMAIL_RE.sub(' ', text)
    head = re.split(r'[;|\t]', without_contacts)[0]
    comma_split = [part.strip() for part in head.split(',') if part.strip()]
    if not comma_split:
        return Name()
    # "Rossi, Mario" — a comma between two name-looking halves means surname first
    surname_comma = (len(comma_split) >= 2
                     and all(_looks_like_name(part) for part in comma_split[:2]))
    head = ' '.join(comma_split[:2]) if surname_comma else comma_split[0]
    head = _TITLE_RE.sub(' ', head)

    tokens = []
    for token in NAME_TOKEN_RE.findall(head):
        if _strip_accents(token).casefold() in NOT_A_NAME:
            break
        tokens.append(token)
        if len(tokens) == 4:
            break
    if len(tokens) < 2:
        return Name()

    if surname_comma:
        return _split_name(tokens, surname_first=True)

    plain = [_strip_accents(token).casefold() for token in tokens]
    shouted = [token.isupper() and len(token) > 1 for token in tokens]
    if any(shouted) and not all(shouted):
        # "ROSSI Mario" / "Mario ROSSI": the capitals are the surname
        first = [token for token, loud in zip(tokens, shouted, strict=True) if not loud]
        last = [_pretty(token) for token, loud in zip(tokens, shouted, strict=True) if loud]
        return Name(' '.join(first), ' '.join(last), ' '.join(tokens))

    known = [name in GIVEN_NAMES for name in plain]
    if known[0] and not known[-1]:
        return _split_name(tokens, surname_first=False)
    if known[-1] and not known[0]:
        return _split_name(tokens, surname_first=True)

    if all(shouted):
        # nothing but capitals: fall back to the Italian list convention, and say
        # it is a convention rather than a reading
        return _split_name(tokens, surname_first=True, certain=False)
    return _split_name(tokens, surname_first=False, certain=False)


def _looks_like_name(part):
    tokens = NAME_TOKEN_RE.findall(_TITLE_RE.sub(' ', part))
    return bool(tokens) and ' '.join(tokens).strip() == part.strip().replace('  ', ' ') \
        and not any(_strip_accents(token).casefold() in NOT_A_NAME for token in tokens)


def find_diet(text):
    """Dietary requirements, including the ones that are a single word."""
    match = DIET_RE.search(text)
    if match and match.group(1).strip():
        return match.group(1).strip(), match.group(0)
    lowered = _strip_accents(text).casefold()
    for diet in STANDALONE_DIETS:
        if re.search(rf'\b{re.escape(diet)}\b', lowered):
            return diet, diet
    return '', ''


def extract_guest(text, *, row_number=0):
    """Read one line of a participant list.

    Records, for every value, the fragment it came from: when a transfer sheet
    is wrong, the office has to be able to see *why* the rule read it that way.
    """
    text = (text or '').strip()
    guest = Guest(row_number=row_number)
    if not text:
        return guest

    email = EMAIL_RE.search(text)
    if email:
        guest.email = email.group(0)
        guest.evidence['email'] = email.group(0)

    # the phone is looked for outside the email, so its digits cannot be mistaken for one
    phone = PHONE_RE.search(EMAIL_RE.sub(' ', text))
    if phone and sum(char.isdigit() for char in phone.group(0)) >= 8:
        guest.phone = phone.group(0).strip()
        guest.evidence['phone'] = guest.phone

    name = find_name(text)
    guest.first_name, guest.last_name = name.first, name.last
    guest.name_order_certain = name.certain
    if name.evidence:
        guest.evidence['name'] = name.evidence

    guest.pax, pax_evidence = find_pax(text)
    if pax_evidence:
        guest.evidence['pax'] = pax_evidence

    guest.arrival, arrival_evidence = _search_time(ARRIVAL_RE, text)
    if arrival_evidence:
        guest.evidence['arrival'] = arrival_evidence
    guest.departure, departure_evidence = _search_time(DEPARTURE_RE, text)
    if departure_evidence:
        guest.evidence['departure'] = departure_evidence

    own = OWN_TRANSPORT_RE.search(text)
    guest.own_transport = own is not None
    if own:
        guest.evidence['own_transport'] = own.group(0)

    guest.lunch = _meal_requested(text, 'pranzo', 'lunch')
    guest.dinner = _meal_requested(text, 'cena', 'dinner')
    guest.diet_notes, diet_evidence = find_diet(text)
    if diet_evidence:
        guest.evidence['diet_notes'] = diet_evidence
    return guest


# --- reading a whole list -----------------------------------------------------------

HEADER_WORDS = frozenset({'nome', 'cognome', 'email', 'telefono', 'pax', 'arrivo', 'partenza'})
#: Rows that state the person is not coming
NOT_ATTENDING_RE = re.compile(r'\b(?:non\s+partecipa|rinunc|disdett|annullat|cancellat|declin)', re.IGNORECASE)


def is_header(text):
    """Whether a row is the spreadsheet's header rather than a person."""
    words = {_strip_accents(word).casefold().strip() for word in re.split(r'[;,\t|]', text or '')}
    return len(words & HEADER_WORDS) >= 2


def read_list(content, filename=''):
    """Turn an uploaded participant list into one line per person.

    A sponsor sends this as a spreadsheet as often as as a pasted email, so a
    spreadsheet row is flattened into the same free text a pasted line is: the
    rules read either shape without a column mapping to maintain.
    """
    lowered = (filename or '').lower()
    if lowered.endswith(('.xlsx', '.xlsm')):
        import io

        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        lines = []
        for row in workbook.active.iter_rows(values_only=True):
            cells = [str(cell).strip() for cell in row if cell not in (None, '')]
            if cells:
                lines.append(', '.join(cells))
        return lines

    from indico_ecm.services.automator import read_document
    text = read_document(content, filename or 'lista.txt')
    return [line.strip() for line in text.splitlines()]


def import_guest_list(rows):
    """Turn the lines of a list into guests, keeping every rejection explained.

    Returns `(guests, rejected)`. A guest needs a name **or** a contact: a row
    with neither cannot be arranged for, and saying so is more useful than
    inventing a placeholder.
    """
    guests = []
    rejected = []
    for index, raw in enumerate(rows, start=1):
        text = (raw or '').strip()
        if not text or len(text) < 3:
            rejected.append(RejectedRow(index, text, 'riga vuota'))
            continue
        if index == 1 and is_header(text):
            continue
        if NOT_ATTENDING_RE.search(text):
            rejected.append(RejectedRow(index, text, 'non partecipa'))
            continue
        guest = extract_guest(text, row_number=index)
        if not guest.full_name and not guest.has_contact:
            rejected.append(RejectedRow(index, text, 'né un nome né un contatto'))
            continue
        guests.append(guest)
    return guests, rejected


# --- what has to be arranged --------------------------------------------------------

@dataclass(frozen=True)
class Categories:
    """The guests split by what the office has to book for them."""

    arrivals: tuple = ()
    departures: tuple = ()
    own_transport: tuple = ()
    no_transfer: tuple = ()
    lunch: tuple = ()
    dinner: tuple = ()

    @property
    def covers(self):
        """Meal covers, counting companions: a table is laid for people, not rows."""
        return {'lunch': sum(guest.pax for guest in self.lunch),
                'dinner': sum(guest.pax for guest in self.dinner)}


def categorize(guests):
    """Split the list into the arrangements it implies.

    A guest with their own transport is a special case whatever their times say:
    booking them a seat wastes it.
    """
    arrivals, departures, own, none = [], [], [], []
    lunch, dinner = [], []
    for guest in guests:
        if guest.own_transport:
            own.append(guest)
        elif guest.arrival is not None:
            arrivals.append(guest)
        elif guest.departure is not None:
            departures.append(guest)
        else:
            none.append(guest)
        if guest.lunch:
            lunch.append(guest)
        if guest.dinner:
            dinner.append(guest)
    return Categories(tuple(arrivals), tuple(departures), tuple(own), tuple(none),
                      tuple(lunch), tuple(dinner))


@dataclass(frozen=True)
class TransferConfig:
    """How the shuttles are organised."""

    #: 'vehicle' fills vehicles to capacity; 'time' only groups by window
    strategy: str = 'vehicle'
    #: minutes
    window: int = 60
    seats_per_vehicle: int = 8

    def __post_init__(self):
        if self.window <= 0:
            raise ValueError('la finestra oraria deve essere positiva')
        if self.seats_per_vehicle <= 0:
            raise ValueError('i posti per veicolo devono essere positivi')


@dataclass(frozen=True)
class TransferGroup:
    """One shuttle run: a time window, the people on it, the seats used."""

    window: str
    guests: tuple
    vehicle_number: int = 1

    @property
    def pax(self):
        return sum(guest.pax for guest in self.guests)


def _window_label(value, minutes):
    start_minutes = (value.hour * 60 + value.minute) // minutes * minutes
    end_minutes = start_minutes + minutes
    start = f'{start_minutes // 60:02d}:{start_minutes % 60:02d}'
    end = f'{end_minutes // 60 % 24:02d}:{end_minutes % 60:02d}'
    return f'{start} - {end}'


def group_transfers(guests, config=None, *, arrival=True):
    """Pack the guests into shuttle runs.

    A party of two never gets split across vehicles: `pax` travels together or
    the arrangement is pointless. A party larger than a vehicle is reported as
    it is instead of being silently truncated — that is a phone call to make,
    not a number to round.
    """
    config = config or TransferConfig()
    attribute = 'arrival' if arrival else 'departure'
    scheduled = [guest for guest in guests
                 if not guest.own_transport and getattr(guest, attribute) is not None]
    scheduled.sort(key=lambda guest: (getattr(guest, attribute), guest.last_name, guest.first_name))

    windows = {}
    for guest in scheduled:
        windows.setdefault(_window_label(getattr(guest, attribute), config.window), []).append(guest)

    groups = []
    for label in sorted(windows):
        members = windows[label]
        if config.strategy != 'vehicle':
            groups.append(TransferGroup(label, tuple(members)))
            continue
        remaining = list(members)
        number = 1
        while remaining:
            load, seats = [], 0
            for guest in list(remaining):
                if load and seats + guest.pax > config.seats_per_vehicle:
                    continue
                load.append(guest)
                seats += guest.pax
                remaining.remove(guest)
                if seats >= config.seats_per_vehicle:
                    break
            groups.append(TransferGroup(label, tuple(load), number))
            number += 1
    return tuple(groups)


def oversized_parties(guests, config=None):
    """Parties that do not fit in one vehicle, which need a call rather than a seat."""
    config = config or TransferConfig()
    return tuple(guest for guest in guests
                 if not guest.own_transport and guest.pax > config.seats_per_vehicle)


def merge_guest(guest, **corrections):
    """A corrected copy, for when a person fixes what the rule read wrong."""
    known = {key: value for key, value in corrections.items() if hasattr(guest, key)}
    return replace(guest, **known)
