# This file is part of the ECM extension for Indico.
# Copyright (C) 2026 - present
#
# Licensed under the MIT License; see the LICENSE file for details.

"""The mail merge, end to end.

Reads the hospital spreadsheet the office already keeps, stores one row per
hospital with its cost sheet, and produces the invitation letters from the
provider's Word template.

The spreadsheet is read by column *label* as well as by key, because the files
in circulation have Italian headers ("Nome Ospedale", "N° Medici") rather than
the internal names.
"""

import csv
import io
import zipfile
from pathlib import Path

from indico.core.db import db
from indico.core.logger import Logger
from indico.util.date_time import now_utc

from indico_ecm.models.operations import InvitationBatch
from indico_ecm.services.costs import cost_sheet_from_legacy
from indico_ecm.services.documents import missing_placeholders, render_docx
from indico_ecm.services.event_schema import INVITATION_FIELDS
from indico_ecm.services.legacy_import import ImportIssue, import_invitation
from indico_ecm.services.letters import InvitationRow, invitation_filename, letter_context


logger = Logger.get('plugin.ecm.invitations')

TEMPLATE_PATH = Path(__file__).parent.parent / 'templates' / 'letters' / 'lettera_invito.docx'

#: label -> key, so a spreadsheet with Italian headers just works
LABEL_TO_KEY = {item.label.strip().casefold(): item.key for item in INVITATION_FIELDS}
KEYS = {item.key for item in INVITATION_FIELDS}


def normalize_headers(headers):
    """Map spreadsheet headers to internal keys, keeping unknown ones out."""
    mapping = {}
    for index, header in enumerate(headers):
        name = (header or '').strip()
        if name in KEYS:
            mapping[index] = name
        elif name.casefold() in LABEL_TO_KEY:
            mapping[index] = LABEL_TO_KEY[name.casefold()]
    return mapping


def read_rows(content, *, filename=''):
    """Read a CSV or XLSX file into dictionaries keyed by internal names."""
    if filename.lower().endswith(('.xlsx', '.xlsm')):
        return _read_xlsx(content)
    return _read_csv(content)


def _read_csv(content):
    text = content.decode('utf-8-sig') if isinstance(content, bytes) else content
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',;\t')
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    try:
        headers = next(reader)
    except StopIteration:
        return []
    mapping = normalize_headers(headers)
    return [{key: row[index] for index, key in mapping.items() if index < len(row)} for row in reader]


def _read_xlsx(content):
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        return []
    mapping = normalize_headers([str(h) if h is not None else '' for h in headers])
    result = []
    for row in rows:
        if not any(cell not in (None, '') for cell in row):
            continue
        result.append({key: row[index] for index, key in mapping.items() if index < len(row)})
    return result


def import_rows(event, rows, *, replace=False):
    """Store mail merge rows against an event.

    Returns `(created, issues)`. Rows that cannot be read are reported, not
    guessed at: an invitation with the wrong number of physicians is a wrong
    letter and a wrong cost.
    """
    issues = []
    if replace:
        InvitationBatch.query.filter_by(event_id=event.id).delete()
    created = []
    for number, raw in enumerate(rows, start=1):
        invitation, _costs = import_invitation(raw, row_number=number, issues=issues)
        if not invitation.hospital:
            issues.append(ImportIssue(number, 'nomeOspedale', 'ospedale mancante'))
            continue
        sheet = cost_sheet_from_legacy(raw, physicians=invitation.physician_count)
        row = InvitationBatch(
            event_id=event.id,
            hospital=invitation.hospital,
            recipient=invitation.recipient,
            recipient_email=str(raw.get('mail') or '').strip(),
            cc_email=str(raw.get('ccMail') or '').strip(),
            department=invitation.department,
            specialty=str(raw.get('specialita') or '').strip(),
            role=invitation.role,
            physician_count=invitation.physician_count,
            sponsor=str(raw.get('sponsor') or '').strip(),
            costs=sheet.as_dict(),
            notes=invitation.notes,
        )
        db.session.add(row)
        created.append(row)
    db.session.flush()
    record_organizations(created, event=event)
    logger.info('imported %d invitation rows for event %d (%d issues)', len(created), event.id, len(issues))
    return created, issues


def record_organizations(rows, *, event=None):
    """Put the hospitals and the sponsor of a mail merge into the CRM.

    These are the organizations the provider actually works with — the sheet is
    where their names arrive — and the CRM already has the two kinds they belong
    to. Without this they stayed on the invitation rows and the companies page
    was empty.

    Only the names are recorded, and the recipient is not turned into a contact:
    a person who ends up on a certificate is created deliberately, not as a side
    effect of importing a spreadsheet.
    """
    try:
        from indico_crm.models.companies import CompanyKind
        from indico_crm.models.links import CRMObjectType, IndicoObjectType, LinkSource, ObjectLink
        from indico_crm.plugin import CRMPlugin
        from indico_crm.services.identity import find_or_create_company
    except ImportError:
        # The CRM plugin is not installed: the invitations still work.
        return []

    if not CRMPlugin.settings.get('autocreate_companies'):
        return []

    recorded = []
    for row in rows:
        for name, kind in ((row.hospital, CompanyKind.healthcare_org),
                           (row.sponsor, CompanyKind.sponsor)):
            company = find_or_create_company(name, kind=kind)
            if company is None:
                continue
            recorded.append(company)
            if event is None:
                continue
            existing = ObjectLink.query.filter_by(
                crm_type=CRMObjectType.company, crm_id=company.id,
                indico_type=IndicoObjectType.event, indico_id=event.id,
                relation=kind.name).first()
            if existing is None:
                db.session.add(ObjectLink(
                    crm_type=CRMObjectType.company, crm_id=company.id,
                    indico_type=IndicoObjectType.event, indico_id=event.id,
                    relation=kind.name, source=LinkSource.import_))
    db.session.flush()
    return recorded


def row_to_invitation(row, *, event=None):
    event = event if event is not None else row.event
    date_text = ''
    if event is not None and event.start_dt:
        start = event.start_dt.strftime('%d/%m/%Y')
        end = event.end_dt.strftime('%d/%m/%Y') if event.end_dt else start
        date_text = start if start == end else f'{start} - {end}'
    return InvitationRow(
        hospital=row.hospital,
        physician_count=row.physician_count,
        role=row.role,
        department=row.department,
        event_name=(event.title if event is not None else ''),
        event_place=(getattr(event, 'venue_name', '') or '' if event is not None else ''),
        date_text=date_text,
        recipient=row.recipient,
        notes=row.notes,
    )


def render_letter(row, *, user_name='', event=None, template_path=TEMPLATE_PATH):
    """Render one invitation letter, returning `(filename, docx_bytes)`."""
    invitation = row_to_invitation(row, event=event)
    context = letter_context(invitation, user_name=user_name)
    content = render_docx(template_path, context)
    row.letter_generated_dt = now_utc()
    return invitation_filename(invitation, extension='docx'), content


def render_batch(event, *, user_name='', template_path=TEMPLATE_PATH):
    """Render every letter of an event into a single zip archive.

    One file per hospital, named the way the office already files them.
    """
    rows = InvitationBatch.query.filter_by(event_id=event.id).order_by(InvitationBatch.hospital).all()
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as archive:
        for row in rows:
            filename, content = render_letter(row, user_name=user_name, event=event,
                                              template_path=template_path)
            archive.writestr(filename, content)
    db.session.flush()
    return len(rows), buffer.getvalue()


def check_template(event, *, template_path=TEMPLATE_PATH):
    """Placeholders the template needs and the data does not provide."""
    row = InvitationBatch.query.filter_by(event_id=event.id).first()
    if row is None:
        return []
    context = letter_context(row_to_invitation(row, event=event))
    return missing_placeholders(template_path, context)
