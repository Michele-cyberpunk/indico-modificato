# This file is part of the ECM extension for Indico.
# Copyright (C) 2026 - present
#
# Licensed under the MIT License; see the LICENSE file for details.

import io
import zipfile
from datetime import date, time

import pytest

from indico_ecm.services.guests import Guest, TransferConfig, group_transfers
from indico_ecm.services.transfer_export import (COLUMNS, SHEET_ITINERARY, SHEET_OUTBOUND, SHEET_RETURN,
                                                 build_workbook, guest_row, itinerary_rows, summary)


@pytest.fixture
def guests():
    return [
        Guest(first_name='Mario', last_name='Rossi', phone='333', pax=2,
              arrival=time(10, 30), departure=time(18, 0), transfer_place='Milano', dinner=True),
        Guest(first_name='Laura', last_name='Bianchi', pax=1, arrival=time(10, 45),
              transfer_place='Roma'),
        Guest(first_name='Anna', last_name='Verdi', pax=1, own_transport=True,
              transfer_place='Torino'),
    ]


def test_the_columns_are_the_ones_the_office_reads():
    assert COLUMNS == ('Cognome', 'Nome', 'Città', 'Mezzo', 'Orario Arrivo', 'Orario Partenza',
                       'Cena', 'Note', 'Data andata', 'Data ritorno')


def test_a_guest_becomes_the_row_the_office_expects(guests):
    row = guest_row(guests[0], outbound_date='15/09/2026', return_date='16/09/2026')
    assert row[:4] == ('Rossi', 'Mario', 'Milano', 'Navetta')
    assert row[4:7] == ('10:30', '18:00', 'Sì')
    assert row[8:] == ('15/09/2026', '16/09/2026')


def test_someone_on_their_own_wheels_is_marked_as_such(guests):
    assert guest_row(guests[2])[3] == 'Proprio'


def test_the_four_counters(guests):
    counts = summary(guests)
    # Two people travel with Rossi, so the head count is not the row count.
    assert counts['OSPITI TOTALI'] == 4
    assert counts['TRASPORTO PROPRIO'] == 1
    assert counts['CENE RICHIESTE'] == 2
    # Only Rossi has both an arrival and a departure.
    assert counts['CON PERNOTTAMENTO'] == 1


def test_an_empty_list_counts_to_zero():
    assert summary([]) == {'OSPITI TOTALI': 0, 'TRASPORTO PROPRIO': 0,
                           'CENE RICHIESTE': 0, 'CON PERNOTTAMENTO': 0}


def test_an_itinerary_is_one_block_per_run(guests):
    groups = group_transfers(guests, TransferConfig(), arrival=True)
    rows = itinerary_rows(groups, arrival=True)
    assert rows[0][0].startswith('Arrivo ')
    assert rows[1] == ('Nome', 'Telefono', 'Pax', 'Orario')
    names = [row[0] for row in rows]
    assert 'Mario Rossi' in names


def test_the_workbook_has_the_three_sheets(guests):
    from openpyxl import load_workbook

    content = build_workbook(guests, event_date=date(2026, 9, 15),
                             arrivals=group_transfers(guests, TransferConfig(), arrival=True),
                             departures=group_transfers(guests, TransferConfig(), arrival=False))
    assert zipfile.is_zipfile(io.BytesIO(content))
    workbook = load_workbook(io.BytesIO(content))
    assert workbook.sheetnames == [SHEET_OUTBOUND, SHEET_RETURN, SHEET_ITINERARY]


def test_the_outbound_sheet_leaves_out_who_drives_themselves(guests):
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(build_workbook(guests, event_date=date(2026, 9, 15))))
    surnames = [row[0] for row in workbook[SHEET_OUTBOUND].iter_rows(min_row=2, values_only=True)]
    assert 'Verdi' not in surnames
    assert 'Rossi' in surnames


def test_without_itineraries_only_two_sheets_are_written(guests):
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(build_workbook(guests)))
    assert workbook.sheetnames == [SHEET_OUTBOUND, SHEET_RETURN]
